"""
Diligent / DTX String Diff Extractor
-------------------------------------
Compares two folder trees (Previous version vs Latest version) and extracts
new, changed, and removed strings across YML, JSON, resjson, and .strings files.
Outputs an RWS-branded Excel with one sheet per subfolder + a summary sheet.

Usage:
    python diligent_string_diff.py <previous_folder> <latest_folder> [output.xlsx]

Example:
    python diligent_string_diff.py "C:/DTX/064_BE/Previous/To Translate" \
                                   "C:/DTX/064_BE/Latest/To Translate" \
                                   064_BE_StringDiff.xlsx
"""

import sys
import os
import json
import re
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── RWS brand colours ────────────────────────────────────────────────────────
RWS_PURPLE = "3E016F"
RWS_PINK   = "E60054"
RWS_TEAL   = "00A89F"
RWS_AMBER  = "FFC700"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F5F5"
MID_GREY   = "E0E0E0"
DARK_TEXT  = "1A1A2E"
GREEN_OK   = "16A34A"
AMBER_WARN = "D97706"
RED_ALERT  = "DC2626"
NOTE_BG    = "F9F4FF"

STATUS_COLOURS = {
    "NEW":     ("E8F5E9", "16A34A"),   # light green fill / green text
    "CHANGED": ("FFF8E1", "D97706"),   # light amber fill / amber text
    "REMOVED": ("FFEBEE", "DC2626"),   # light red fill / red text
}

thin  = Side(style="thin",   color="D0D0D0")
thick = Side(style="medium", color="D0D0D0")
thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

def _fill(hex_): return PatternFill("solid", fgColor=hex_)
def _font(bold=False, size=10, color=DARK_TEXT, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── File parsers ──────────────────────────────────────────────────────────────

def parse_json(path: Path) -> dict:
    """Parse flat or nested JSON/resjson into key→value dict."""
    try:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        data = json.loads(text)
    except Exception:
        return {}
    return _flatten(data)

def _flatten(obj, prefix=""):
    items = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            full_key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, (dict, list)):
                items.update(_flatten(v, full_key))
            else:
                items[full_key] = str(v) if v is not None else ""
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            items.update(_flatten(v, f"{prefix}[{i}]"))
    return items

def parse_yaml(path: Path) -> dict:
    """Parse YAML using PyYAML if available, else basic key: value."""
    try:
        import yaml
        with path.open(encoding="utf-8-sig", errors="replace") as f:
            data = yaml.safe_load(f)
        return _flatten(data) if isinstance(data, (dict, list)) else {}
    except ImportError:
        pass
    # Fallback: simple key: "value" regex
    result = {}
    pattern = re.compile(r'^(\s*)([^#\s][^:]*?):\s*["\']?(.*?)["\']?\s*$')
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        m = pattern.match(line)
        if m:
            key = m.group(2).strip()
            val = m.group(3).strip().strip('"\'')
            if key and val:
                result[key] = val
    return result

def parse_strings(path: Path) -> dict:
    """Parse Apple .strings format: "key" = "value";"""
    result = {}
    pattern = re.compile(r'^"(.+?)"\s*=\s*"(.*?)"\s*;')
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        m = pattern.match(line.strip())
        if m:
            result[m.group(1)] = m.group(2)
    return result

PARSERS = {
    ".json":    parse_json,
    ".resjson": parse_json,
    ".yml":     parse_yaml,
    ".yaml":    parse_yaml,
    ".strings": parse_strings,
}

def parse_file(path: Path) -> dict:
    parser = PARSERS.get(path.suffix.lower())
    return parser(path) if parser else {}

# ── Folder walker ─────────────────────────────────────────────────────────────

def collect_files(root: Path) -> dict[str, Path]:
    """Return {relative_posix_path: absolute_path} for all supported files."""
    files = {}
    for p in root.rglob("*"):
        if p.suffix.lower() in PARSERS and p.is_file():
            files[p.relative_to(root).as_posix()] = p
    return files

# ── Diff engine ───────────────────────────────────────────────────────────────

def diff_folders(prev_root: Path, new_root: Path):
    """
    Returns list of dicts:
    {subfolder, file, key, status, old_value, new_value}
    """
    prev_files = collect_files(prev_root)
    new_files  = collect_files(new_root)

    all_rel = sorted(set(prev_files) | set(new_files))
    rows = []

    for rel in all_rel:
        prev_kv = parse_file(prev_files[rel]) if rel in prev_files else {}
        new_kv  = parse_file(new_files[rel])  if rel in new_files  else {}

        all_keys = sorted(set(prev_kv) | set(new_kv))
        for key in all_keys:
            old_val = prev_kv.get(key)
            new_val = new_kv.get(key)

            if old_val is None and new_val is not None:
                status = "NEW"
            elif old_val is not None and new_val is None:
                status = "REMOVED"
            elif old_val != new_val:
                status = "CHANGED"
            else:
                continue  # identical — skip

            parts = rel.split("/")
            subfolder = parts[0] if len(parts) > 1 else "(root)"
            rows.append({
                "subfolder": subfolder,
                "file":      rel,
                "key":       key,
                "status":    status,
                "old_value": old_val or "",
                "new_value": new_val or "",
            })

    return rows

# ── Excel builder ─────────────────────────────────────────────────────────────

HEADERS = ["#", "File", "String Key", "Status", "Previous Value", "New / Current Value"]
COL_WIDTHS = [5, 40, 40, 12, 50, 50]

def write_title(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=13, color=WHITE)
    c.fill      = _fill(RWS_PURPLE)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 26

def write_note(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(size=9, color="666666")
    c.fill      = _fill(NOTE_BG)
    c.alignment = _align("left", wrap=True)
    ws.row_dimensions[row].height = 32

def write_headers(ws, row, labels):
    for i, h in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _font(bold=True, size=10, color=WHITE)
        c.fill      = _fill(RWS_PURPLE)
        c.alignment = _align("center", wrap=True)
        c.border    = thin_b
    ws.row_dimensions[row].height = 34

def write_data_row(ws, row, values, status, alt):
    fill_hex, text_hex = STATUS_COLOURS.get(status, (WHITE, DARK_TEXT))
    row_fill = fill_hex if status in STATUS_COLOURS else (LIGHT_GREY if alt else WHITE)

    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font      = _font(size=10, color=text_hex if i == 4 else DARK_TEXT,
                            bold=(i == 4))
        c.fill      = _fill(row_fill)
        c.alignment = _align("left", wrap=True)
        c.border    = thin_b
    ws.row_dimensions[row].height = 30

def write_sheet(ws, title, note, rows_data, prev_label, new_label):
    n = len(HEADERS)
    write_title(ws, 1, title, n)
    write_note(ws, 2, note, n)
    ws.row_dimensions[3].height = 8   # spacer
    write_headers(ws, 4, HEADERS)
    ws.freeze_panes = "A5"

    for i, r in enumerate(rows_data):
        excel_row = 5 + i
        values = [
            i + 1,
            r["file"],
            r["key"],
            r["status"],
            r["old_value"],
            r["new_value"],
        ]
        write_data_row(ws, excel_row, values, r["status"], alt=(i % 2 == 1))

    # Column widths
    for col_i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # Tab colour
    ws.sheet_properties.tabColor = RWS_TEAL

def write_summary_sheet(ws, all_rows, prev_label, new_label):
    by_sub = defaultdict(lambda: defaultdict(int))
    for r in all_rows:
        by_sub[r["subfolder"]][r["status"]] += 1

    n = 6
    write_title(ws, 1, "String Diff Summary", n)
    write_note(ws, 2,
               f"Comparing: {prev_label}  →  {new_label}  |  "
               f"Total changes: {len(all_rows)}", n)
    ws.row_dimensions[3].height = 8

    hdrs = ["Subfolder", "NEW Strings", "CHANGED Strings", "REMOVED Strings", "Total Changes"]
    write_headers(ws, 4, hdrs)
    ws.freeze_panes = "A5"

    col_widths_s = [30, 18, 20, 20, 16]
    for col_i, w in enumerate(col_widths_s, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    total_new = total_chg = total_rem = 0
    for i, (sub, counts) in enumerate(sorted(by_sub.items())):
        row = 5 + i
        new_ = counts.get("NEW", 0)
        chg_ = counts.get("CHANGED", 0)
        rem_ = counts.get("REMOVED", 0)
        tot_ = new_ + chg_ + rem_
        total_new += new_; total_chg += chg_; total_rem += rem_
        alt = (i % 2 == 1)
        vals = [sub, new_, chg_, rem_, tot_]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font      = _font(size=10)
            c.fill      = _fill(LIGHT_GREY if alt else WHITE)
            c.alignment = _align("left" if ci == 1 else "center")
            c.border    = thin_b

    # Grand total
    gt_row = 5 + len(by_sub)
    gt_vals = ["TOTAL", total_new, total_chg, total_rem, total_new + total_chg + total_rem]
    pink_s  = Side(style="medium", color=RWS_PINK)
    gt_border = Border(left=thin, right=thin, top=pink_s, bottom=pink_s)
    for ci, v in enumerate(gt_vals, 1):
        c = ws.cell(row=gt_row, column=ci, value=v)
        c.font      = _font(bold=True, size=10, color=RWS_PINK)
        c.fill      = _fill(MID_GREY)
        c.border    = gt_border
        c.alignment = _align("left" if ci == 1 else "center")

    ws.sheet_properties.tabColor = RWS_PURPLE

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    prev_root = Path(sys.argv[1])
    new_root  = Path(sys.argv[2])
    output    = Path(sys.argv[3]) if len(sys.argv) > 3 else Path("StringDiff_Output.xlsx")

    if not prev_root.exists():
        print(f"ERROR: Previous folder not found: {prev_root}")
        sys.exit(1)
    if not new_root.exists():
        print(f"ERROR: Latest folder not found: {new_root}")
        sys.exit(1)

    print(f"Scanning previous: {prev_root}")
    print(f"Scanning latest:   {new_root}")

    all_rows = diff_folders(prev_root, new_root)

    if not all_rows:
        print("No differences found between the two folders.")
        sys.exit(0)

    print(f"Found {len(all_rows)} differences. Building Excel…")

    by_sub = defaultdict(list)
    for r in all_rows:
        by_sub[r["subfolder"]].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    prev_label = prev_root.name
    new_label  = new_root.name

    # Summary sheet first
    ws_sum = wb.create_sheet("Summary")
    write_summary_sheet(ws_sum, all_rows, prev_label, new_label)

    # One sheet per subfolder
    for sub in sorted(by_sub):
        safe_name = sub[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(safe_name)
        title = f"String Changes — {sub}"
        note  = (f"Previous: {prev_label}   |   Latest: {new_label}   |   "
                 f"{len(by_sub[sub])} change(s)")
        write_sheet(ws, title, note, by_sub[sub], prev_label, new_label)

    # All-changes sheet
    ws_all = wb.create_sheet("All Changes")
    write_sheet(ws_all, "All String Changes",
                f"Previous: {prev_label}   |   Latest: {new_label}   |   "
                f"{len(all_rows)} total change(s)",
                all_rows, prev_label, new_label)
    ws_all.sheet_properties.tabColor = RWS_PINK

    wb.save(output)
    print(f"Saved: {output}")
    return str(output)

if __name__ == "__main__":
    main()
